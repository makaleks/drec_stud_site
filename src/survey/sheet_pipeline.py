# coding: utf-8

# Usage: 
# Create your own function with:
#   [in]  - previous list of dictionary result
#   [out] - list to fill an Excel line
# Then, put the function name into list at the end of this file
# The 'name' in the pipeline defines how it will be displayed in the web (future)
# The order in the list defines the order of execution
# Default [in] is ['is_anonymous': True/False, 'answ_data': [...], lines: [{['header'], {'structure'}, ['results']}]]
# 'structure' and 'answ_data' are as defined in 'models.py'
# The default pipeline adds ['total', 'percent']

# Returns {'author': None/User, 'question1': [value], 'question2': ...}
# Each 'name' (aka 'id') is converted to 'title'

import json
import openpyxl

def _get_answ_len(question):
    return len(question['structure']['choices'])

def _get_data_from_raw(list_with_structures, answ_data_single):
    result = {'author': None}
    data = json.loads(answ_data_single.value)
    for l in list_with_structures:
        s = l['structure']
        if answ_data_single.answer:
            result['author'] = answ_data_single.answer.user
            # If no choices, it is a text
        if s['choices']:
            if type(data[s['name']]) is list:
                lst = []
                for l in data[s['name']]:
                    for choice in s['choices']:
                        if choice['value'] == l:
                            lst.append(choice['text'])
                            break
                result[s['title']] = lst
            else:
                for choice in s['choices']:
                    if choice['value'] == data[s['name']]:
                        # Wrapped in list for universal access in future
                        result[s['title']] = [choice['text']]
                        break
    return result

# Warning: if you change this function name, make sure 
# that 'percent()' calls the new name
def total(source):
    answ_data = source['answ_data']
    lines = source['lines']
    #print(lines)
    #print('###')
    # Fill with zeros
    for question in lines:
        question['results'].append([0]*_get_answ_len(question))
    for d in answ_data:
        dic = _get_data_from_raw(lines, d)
        #print(dic)
        #print('### question')
        for question in lines:
            #print(question)
            #print('### list')
            # All chosen answers in list, stored in dictionary
            for answ in dic[question['structure']['title']]:
                #print('# Index for \'{0}\' is {1}'.format(answ, question['header'].index(answ) - 1))
                question['results'][-1][question['header'].index(answ) - 1] += 1
    for question in lines:
        in_total = 0
        for v in question['results'][-1]:
            in_total += v
        question['results'][-1].append(in_total)

def percent(source):
    for question in source['lines']:
        vals = question['results'][source['pipe_order'].index('total')]
        in_total = 0
        for v in vals[:_get_answ_len(question)]:
            in_total += v
        question['results'].append([0]*_get_answ_len(question))
        for i in range(_get_answ_len(question)):
            val = (vals[i]*100) / in_total
            question['results'][-1][i] = val
        question['results'][-1].append(100)

def form(wb, pipeline_order,lines, possible_colors):
    ws = wb.active
    ws.column_dimensions['A'].width = 30
    # Columns
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[openpyxl.utils.cell.get_column_letter(i)].width = 25
    # Rows
    for i in range(len(lines)):
        row = ws.max_row - (i + 1)*(len(pipeline_order) + 1) + 1
        # height
        ws.row_dimensions[row].height = 30
        # alignment
        ws['A' + str(row)].alignment = openpyxl.styles.Alignment(wrapText = True)
        # color on first column
        color = openpyxl.styles.PatternFill(fill_type = 'solid', start_color = possible_colors[i % len(possible_colors)], end_color = possible_colors[i % len(possible_colors)])
        for j in range(len(pipeline_order)):
            pos = 'A' + str(row + j + 1)
            ws[pos].fill = color
            for side in ['left','right','top','bottom']:
                getattr(ws[pos].border, side).border_style = 'thin'
        # color on top row
        for c in range(1, ws.max_column + 1):
            pos = openpyxl.utils.cell.get_column_letter(c) + str(row)
            ws[pos].fill = color
            for side in ['left','right','top','bottom']:
                getattr(ws[pos].border, side).border_style = 'thin'

pipeline_order = [{'Всего': total}, {'Проценты': percent}]
