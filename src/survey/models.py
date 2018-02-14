# coding: utf-8
from django.db import models
from django.conf import settings
from django.urls import reverse
from django.utils import timezone
import json
import datetime
import openpyxl

from .sheet_pipeline import pipeline_order

# Create your models here.

class Survey(models.Model):
    title           = models.CharField(max_length = 64, blank = False, null = False, verbose_name    = 'Заголовок')
    description = models.CharField(max_length = 256, blank = False, null = False, verbose_name    = 'Описание')
    structure       = models.TextField(blank = False, null = False, verbose_name = 'Опрос в JSON (survey.js)')
    # Surveys can still be edited
    started         = models.DateTimeField(blank = False, null = True, verbose_name = 'Дата начала')
    finished        = models.DateTimeField(blank = False, null = True, verbose_name = 'Дата конца')
    is_anonymous    = models.BooleanField(default = False, null = False, verbose_name = 'Анонимный (навсегда!))')
    allow_rewrite   = models.BooleanField(default = True, null = False, verbose_name = 'Разрешить исправлять')
    sheet           = models.FileField(upload_to='surveys', blank = True, null = True, verbose_name = 'Свежий отчёт')
    def __str__(self):
        return self.title
    def is_started():
        return started and started > timezone.now()
    def is_finished():
        return started and finished
    def start(self, date_start = timezone.now(), date_end = None):
        self.started = start
        self.finished = end
        self.save(update_fields=['started', 'finished'])
    def get_absolute_url(self):
        return reverse('survey_detail', args = (self.pk,))
    def get_edit_url(self):
        return reverse('survey_edit', args=(self.pk,))
    def save(self, *args, **kwargs):
        if self.is_anonymous:
            self.allow_rewrite = False
        return super(Survey, self).save(*args, **kwargs)
    # gen list of RGB (without WHITE)
    def gen_colors(step, start):
        result = []
        r = start
        while r <= 255:
            g = start
            while g <= 255:
                b = start
                while b <= 255:
                    result.append('{:02X}{:02X}{:02X}'.format(r,g,b))
                    b += step
                g += step
            r += step
        # delete BLACK
        result.pop(0)
        # delete WHITE
        result.pop()
        return result
    def get_structure(self):
        # recursive search to extract survey information of same structure
        # returns [{ 'name', 'title', 'choices':[{'value','text'}], 'group_name'}, 'type'], group_names (all that were found)
        # legacy note: group names are not used in final sheet for now
        def _process(source, result, group_name, groups):
            # fill result, usually this function is recursive
            if type(source) is list:
                for s in source:
                    _process(s, result, group_name, groups)
            elif source['type'] == 'panel':
                group = source.get('title')
                if not group:
                    group = source['name']
                if not group in groups:
                    groups.append(group)
                for e in source['elements']:
                    _process(e, result, group, groups)
            else:
                to_append = {}
                to_append['name'] = source['name']
                to_append['type'] = source['type']
                to_append['choices'] = []
                to_append['title'] = source['title'] if 'title' in source else source['name']
                if source['type'] != 'text':
                    if source['type'] == 'rating':
                        if 'rateValues' not in source:
                            to_append['choices'] = [{'text': str(i), 'value': str(i)} for i in range(1,6)]
                        else:
                            to_append['choices'] = source['rateValues']
                            for i in range(len(to_append['choices'])):
                                to_append['choices'][i] = {'text': to_append['choices'][i], 'value': to_append['choices'][i]}
                    else:
                        to_append['choices'] = source['choices']
                    for c in to_append['choices']:
                        if not 'text' in c:
                            c['text'] = c['name']
                to_append['group_name'] = group_name
                if not to_append:
                    print('# found!\n  {0}'.format(source))
                result.append(to_append)
        source = json.loads(self.structure)
        pages = source['pages']
        result = []
        groups = []
        is_first = True
        for p in pages:
            _process(p['elements'], result, '', groups)
            is_first = False
        return result, groups
    def gen_sheet(self):
        def _get_pos(col, row):
            return openpyxl.utils.cell.get_column_letter(col) + str(row)
        # colors: 0, 170, 170+85 
        # => [3 : 0,170,255] ^ [3 : rgb] - [2 : rgb(0,0,0)] = 26 colors to use
        possible_colors = Survey.gen_colors(step = 85, start = 170)
        layout, groups = self.get_structure()
        group_to_color = {groups[i]: possible_colors[i%len(groups)] for i in range(len(groups))}
        # WHITE - default background
        group_to_color[''] = 'FFFFFF'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Числа'
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['A'].wrapText = True
        ws.append(['Опрос:', self.title])
        ws.append(['Анонимный', 'Да' if self.is_anonymous else 'Нет'])
        ws.append(['Начало опроса', self.started.date()])
        ws.append(['Окончание опроса', self.finished.date()])
        if datetime.datetime.now() < self.finished:
            ws.append(['Дней осталось', (self.finished - datetime.datetime.now()).days])
        ws.append(['Ответов собрано', self.answers.count()])
        ws.append(['Последнее обновление', timezone.now()])

        answers = {}

        answ_data = list(self.answers_data.all())

        # Color index
        col = 0
        # All lines will be in memory first
        lines = []
        # Generate all title lines
        for question in layout:
            if len(question['choices']):
                line = [question['title']]
                line.extend([choice['text'] for choice in question['choices']])
                lines.append({'header': line, 'structure': question, 'results': []})
        source_data = {'is_anonymous': self.is_anonymous, 'answ_structure': layout, 'answ_data': answ_data, 'lines': lines, 'pipe_order': [next(iter(n.values())).__name__ for n in pipeline_order]}
        for p in pipeline_order:
            f = next(iter(p.values()))
            f(source_data)
        # Start normalize len (num of columns)
        # |yes| no|add|   | => |yes| no|   |add|
        # | 1 | 2 | 3 |add| => | 1 | 2 | 3 |add|
        max_len = 0
        for question in layout:
            l = len(question['choices'])
            if l > max_len:
                max_len = l
        for i in range(len(lines)):
            # -1 is for 'question' column
            l = len(lines[i]['header']) - 1
            if l < max_len:
                for j in range(len(pipeline_order)):
                    # Insert [i:i] syntax for case, when it is wider
                    r = lines[i]['results']
                    r[j][l : l] = [None]*(max_len - l)
        # End normalize
        for line in lines:
            ws.append(line['header'])
            for i in range(len(line['results'])):
                # List concatenation
                ws.append([next(iter(pipeline_order[i]))] + line['results'][i])

        # Start Formatting
        ws = wb.active
        ws.column_dimensions['A'].width = 30
        # Columns
        for i in range(2, ws.max_column + 1):
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(i)].width = 15
        # Rows
        for i in range(len(lines)):
            row = ws.max_row - (i + 1)*(len(pipeline_order) + 1) + 1
            # height
            ws.row_dimensions[row].height = 30
            # alignment
            ws['A' + str(row)].alignment = openpyxl.styles.Alignment(wrapText = True)
            # color on first column
            color = openpyxl.styles.PatternFill(fill_type = 'solid', start_color = possible_colors[i % len(possible_colors)], end_color = possible_colors[i % len(possible_colors)])
            for j in range(len(pipeline_order)):
                pos = 'A' + str(row + j + 1)
                ws[pos].fill = color
                for side in ['left','right','top','bottom']:
                    getattr(ws[pos].border, side).border_style = 'thin'
            # color and align on top row
            for c in range(1, ws.max_column + 1):
                ws[pos].alignment = openpyxl.styles.Alignment(wrapText = True)
                pos = openpyxl.utils.cell.get_column_letter(c) + str(row)
                ws[pos].fill = color
                for side in ['left','right','top','bottom']:
                    getattr(ws[pos].border, side).border_style = 'thin'
        # End formatting
        # Start text answers
        ws = wb.create_sheet('Текст')
        col = 1 if self.is_anonymous else 2
        for i in range(len(layout)):
            row = 1
            # Text only
            if layout[i]['type'] == 'text':
                # Header
                color = openpyxl.styles.PatternFill(fill_type = 'solid', start_color = possible_colors[col % len(possible_colors)], end_color = possible_colors[col % len(possible_colors)])
                pos = _get_pos(col, row)
                ws[pos].value = layout[i]['title']
                ws[pos].fill = color
                for side in ['left','right','top','bottom']:
                    getattr(ws[pos].border, side).border_style = 'thin'
                row += 1
                # Answers
                for answ_data_single in answ_data:
                    if not self.is_anonymous and answ_data_single.answer:
                        pos = _get_pos(col - 1, row)
                        ws[pos].value = answ_data_single.answer.user.get_full_name() + str(answ_data_single.answer.user.group_number)
                        ws[pos].fill = color
                        for side in ['left','right','top','bottom']:
                            getattr(ws[pos].border, side).border_style = 'thin'
                    pos = _get_pos(col, row)
                    ws[pos].value = json.loads(answ_data_single.value)[layout[i]['name']]
                    ws[pos].alignment = openpyxl.styles.Alignment(wrapText = True)
                    ws.row_dimensions[row].height = 30
                    row += 1
                col += 1 if self.is_anonymous else 2
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(i)].width = 30
        # End text answers
        return wb
    class Meta:
        verbose_name        = 'Опрос'
        verbose_name_plural = 'Опросы'
        ordering            = ['-started']

class Answer(models.Model):
    survey  = models.ForeignKey(Survey, on_delete = models.CASCADE, related_name = 'answers', verbose_name = 'Опрос')
    user    = models.ForeignKey(settings.AUTH_USER_MODEL, blank = False, null = False, on_delete = models.CASCADE, related_name = 'answers', verbose_name = 'Пользователь')
    created = models.DateTimeField(auto_now = True, blank = False, null = False, verbose_name = 'Дата ответа')
    def __str__(self):
        return 'Ответ на {0}'.format(self.survey.title)
    class Meta:
        verbose_name        = 'Ответ'
        verbose_name_plural = 'Ответы'
        ordering            = ['-created']

class AnswerData(models.Model):
    answer = models.ForeignKey(Answer, default = None, blank = True, null = True, on_delete = models.CASCADE, related_name = 'answer_data', verbose_name = 'Ответ')
    survey  = models.ForeignKey(Survey, on_delete = models.CASCADE, related_name = 'answers_data', verbose_name = 'Опрос')
    value  = models.TextField(blank = False, null = False, verbose_name = 'ответ в JSON (survey.js)')
    class Meta:
        verbose_name        = 'Результат опроса'
        verbose_name_plural = 'Результаты опросов'

