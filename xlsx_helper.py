#!/usr/bin/env python3

import sys, argparse, os, datetime, glob, getpass
from os.path import isfile
import re
from openpyxl import load_workbook
import vk
import json
import progressbar
from requests.exceptions import ReadTimeout

import datetime

_todo_str = 'user_todo_{0}.csv'.format(datetime.datetime.now().strftime('%Y_%m_%d_%H_%M_%S'))
_err_str  = 'user_errors.csv'

from src.utils import validators

_err_countdown = 20
def err_countdown():
    global _err_countdown
    _err_countdown -= 1
    if not _err_countdown:
        sys.exit('Too many errors! Interrupted')

# TODO: v - vertical
# f - first name
# l - last name
# p - patronymic name
# n - full name
# i - account id
# g - group
# t - telephone
# e - email
def filter_format(s, pattern=re.compile(r'^[flpnigte]+$')):
    if s and not pattern.match(s):
        raise argparse.ArgumentTypeError('Invalid format!')
    return s
def filter_file(s):
    if not os.path.exists(s):
        raise argparse.ArgumentTypeError('File \'{0}\' not found!'.format(s))
    return s
def _get_cell_separation_pos(s):
    i = 0
    for c in s:
        if c.isdigit():
            break
        i+=1
    return i
def inc_cell_row(start):
    i = _get_cell_separation_pos(start)
    num = int(start[i:])
    st = start[:i]
    return st + str(num + 1)
def inc_cell_column(start):
    i = _get_cell_separation_pos(start)
    st = start[:i]
    end = start[i:]
    if len(st) == 1 and st != 'Z':
        st = chr(ord(st) + 1)
    else:
        sys.exit('FATAL: inc_cell_columnt is after \'Z\'!')
    return st + end

def add_error_row(field, start, ws):
    s = start
    f = open(_err_str, 'a')
    print('Error:', end=' ')
    while ws[s].value:
        f.write('{0},'.format(ws[s].value))
        print(ws[s].value, end=',')
        s = inc_cell_column(s)
    print('\n')
    f.write(' ({0})\n'.format(field))
    f.close()
    if _err_countdown == 1:
        print('Terminated at {0}'.format(start))
    err_countdown()

_total_todo = 0

def add_todo_row(field, start, ws):
    global _total_todo
    s = start
    f = open(_todo_str, 'a')
    _total_todo += 1
    while ws[s].value:
        f.write('{0},'.format(ws[s].value))
        s = inc_cell_column(s)
    f.write(' ({0})\n'.format(field))
    f.close()

def get_schema_by_start(start, input_format):
    if not input_format:
        # Put some analysis here
        return 'flpigte'
    else:
        return input_format

group_cache = ''

def get_line_by_schema(start, input_format, ws, vk_api):
    global group_cache
    s = start
    result = {}
    for c in input_format:
        if c != input_format[0]:
            s = inc_cell_column(s)
        tmp = str(ws[s].value).strip() if ws[s].value is not None else ''
        if tmp and tmp[0] == '=':
            tmp = tmp[1:]
        tmp = tmp.strip()
        if c == 'n':
            tmp = tmp.split()
            # TODO: normal reaction on No Patconymic
            if len(tmp) != 3 and len(tmp) != 2:
                print('\'{0}\''.format(tmp))
                print('full name check 1: ')
                add_error_row('full name', start, ws)
                return None
            elif len(tmp) == 2:
                add_todo_row('patronymic name', start, ws)
            elif len(tmp) != 3:
                print('\'{0}\''.format(tmp))
                print('full name check 2: ')
                add_error_row('full name', start, ws)
                return None
            for elem in tmp:
                if not validators.is_valid_name(str(elem)):
                    print('\'{0}\''.format(tmp))
                    print('full name check 3: ')
                    add_error_row('full name', start, ws)
                    tmp = False
                    break
            if not tmp:
                return None
            result['last_name'] = tmp[0]
            result['first_name'] = tmp[1]
            # TODO: normal reaction on No Patronymic
            if len(tmp) == 3:
                result['patronymic_name'] = tmp[2] 
            elif len(tmp) == 2:
                result['patronymic_name'] = ''
        elif c in 'lfp':
            if  not validators.is_valid_name(str(tmp)) and not (c == 'p' and not tmp):
                print('\'{0}\''.format(tmp))
                print('name check 1: ')
                add_error_row('first/last/patronymic name', start, ws)
                return None
            elif c == 'p' and not tmp:
                result['patronymic_name'] = ''
            elif c == 'p':
                result['patronymic_name'] = tmp
            elif c == 'l':
                result['last_name'] = tmp
            elif c == 'f':
                result['first_name'] = tmp
            else:
                sys.exit('Unexpected state at \'elif c in \'lfp\'\'')
        elif c == 'i':
            id_is_raw = re.compile(r"^(i|I)d[0-9]+$")
            if not validators.is_valid_id_url(str(tmp)):
                print('\'{0}\''.format(tmp))
                print('vk id check 1: ')
                add_error_row('vk id', start, ws)
                return None
            if tmp[:7] == 'http://':
                tmp = tmp[7:]
            elif tmp[:8] == 'https://':
                tmp = tmp[8:]
            if tmp[:7] == 'vk.com/':
                tmp = tmp[7:]
            elif tmp[:9] == 'm.vk.com/':
                tmp = tmp[9:]
            if (tmp[:2] == 'id' or tmp[:2] == 'Id') and id_is_raw.match(tmp):
                tmp = tmp[2:]
            if tmp[-1] == '/':
                tmp = tmp[:-1]
            flag = True
            while flag:
                try:
                    tmp = vk_api.users.get(user_ids=tmp)
                    #print(str(tmp))
                    result['account_id'] = tmp[0]['uid']
                except BaseException as e:
                    print('{0}\n'.format(e.__class__.__name__))
                    if not isinstance(e, ReadTimeout):
                        flag = False
                        print('{0}\n'.format(vars(e)))
                        print('{0}'.format(str(e)))
                        print('\'{0}\''.format(tmp))
                        print('vk id check 2: ')
                        add_error_row('vk id', start, ws)
                        return None
                    else:
                        print('Time out. Trying again...\n\n')
                        continue
                flag = False
        elif c == 'g':
            if not validators.is_valid_group(str(tmp)) and not group_cache:
                print('\'{0}\''.format(tmp))
                print('group check 1: ')
                add_error_row('group number', start, ws)
                return None
            elif not validators.is_valid_group(str(tmp)):
                tmp = group_cache
            group_cache = tmp
            result['group_number'] = tmp
        elif c == 't':
            if not validators.is_valid_phone(str(tmp)):
                tmp = ''
                add_todo_row('phone number', start, ws)
                #print('\'{0}\''.format(tmp))
                #print('telephone check 1: ')
                #add_error_row('phone number', start, ws)
                #return None
            result['phone_number'] = tmp
        elif c == 'e':
            # I saw so strange emails
            n = tmp.find('@')
            if n != -1:
                tmp = tmp[:n] + tmp[n:].lower()
            # email is not required
            if not validators.is_valid_email(str(tmp)) and tmp:
                tmp = ''
                add_todo_row('phone number', start, ws)
                #print('\'{0}\''.format(tmp))
                #print('email check 1: ')
                #add_error_row('email', start, ws)
                #return None
            result['email'] = tmp
        else:
            sys.exit('FATAL: get_line_by_schema type out of range!')
    return result


def main(argv):
    global _err_str
    global _todo_str
    global _total_todo
    filename = 'drec_stud_site_backup_{0}'.format(datetime.datetime.now().strftime('%Y_%m_%d_%H_%M_%S'))

    parser = argparse.ArgumentParser(description='Fill database with users according to .xlsx or /*.csv*/ file.')

    parser.add_argument('-f', '--format', type=filter_format, default='', required=False, help='set format of file')
    parser.add_argument('-s', '--start', type=str, default='', required=False, help='set starting cell')
    parser.add_argument('filename', type=filter_file, help='Input file name')

    args = parser.parse_args()

    if args.filename[-5:] != '.xlsx' :#and args.filename[-4:] != '.csv':
        sys.exit('Only .xlsx files are supported!')
        #sys.exit('Only .xlsx and .csv files are supported!')
    f = open(_err_str, 'w')
    f.close()
    f = open(_todo_str, 'w')
    f.close()

    wb = load_workbook(args.filename, read_only=True)
    if not wb:
        sys.exit('The library can`t read xlsx!')
    ws = wb.active
    if not ws:
        sys.exit('The xlsx library error on selecting active workbook!')
    if args.start:
        try:
            print('Start value is: {0}\n'.format(ws[args.start].value))
        except:
            sys.exit('Invalid starting cell!')
    start = args.start if args.start else 'A1'
    
    schema = get_schema_by_start(start, args.format)
    vk_api = vk.API(vk.Session())
    result = []
    s = start
    line_dic = get_line_by_schema(start, schema, ws, vk_api)
    print('First row is:   {0}\n'.format(str(line_dic)))

    status = ''
    while status != 'N' and status != 'y':
        print ('Look at first line and accept it was parsed well')
        status = input('[y]Yes or [N]No: ')
        if status == '':
            status = 'N'
        if status != 'y' and status != 'N':
            print('Try again...')
    if status == 'N':
        sys.exit('Result was not accepted by user...')
    result.append(line_dic)
    s = inc_cell_row(s)

    #print('{0}'.format(ws.max_row))
    #print('{0}'.format(int(start[_get_cell_separation_pos(s):])))

    progress_widgets = [progressbar.Percentage(), progressbar.Bar()]
    bar = progressbar.ProgressBar(widgets=progress_widgets, max_value=ws.max_row).start()
    bar.update(0)
    
    while ws[s].value:
        line_dic = get_line_by_schema(s, schema, ws, vk_api)
        if line_dic:
            result.append(line_dic)
        bar.update(int(start[_get_cell_separation_pos(s):]))
        s = inc_cell_row(s)
    bar.finish()

    with open('users_parsed.txt', 'w') as outfile:
        json.dump(result, outfile, indent=2)

    #print ('Trying to process:\n$ ', command_str.replace('\n', '\n$ '))
    #os.system(command_str)
    print('Done!\n')
    if _total_todo != 0:
        print('TODO: fill {0} users (see \'users_todo.csv\' file\n'.format(_total_todo))
    else:
        os.remove(_todo_str)

if __name__ == '__main__':
    main(sys.argv[1:])
